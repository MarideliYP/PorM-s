import os

from django.contrib.contenttypes.models import ContentType
import openpyxl
from django.core.exceptions import PermissionDenied
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.views import View
from .models import Producto, Inmueble, Oferta, Evento, Reserva, CarroCompra, ArticuloCarro, Valoracion, Contrato, \
    ContratoFirmado, Empresa, Media
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy
from datetime import datetime, timedelta
from django.contrib import messages
from django.utils.decorators import method_decorator
from .forms import Producto_form, Inmueble_form, Oferta_form, Evento_form, Reserva_form, Empresa_form, \
    MediaForm
from openpyxl import load_workbook
from django.views.generic import (
    UpdateView,
    CreateView,
    DeleteView,
    ListView, TemplateView,
)


def es_empresa(user):
    return user.groups.filter(name='Empresa').exists()


def es_admin(user):
    return user.groups.filter(name='Administrador').exists()


def es_cliente(user):
    return user.groups.filter(name='Cliente').exists()


def es_repartidor(user):
    return user.groups.filter(name='Repartidor').exists()


def es_admin_o_cliente(user):
    return user.groups.filter(name__in=['Administrador', 'Cliente']).exists()


def es_flotante(user):
    return user.groups.filter(name='Cliente').exists()


def vistaprincipal(request):
    # Verificar roles del usuario (solo si existe en la DB actual)
    is_administrador = False
    is_cliente = False
    is_repartidor = False
    is_empresa = False

    if request.user.is_authenticated:
        try:
            # Forzar carga del usuario desde la DB actual
            request.user.groups.exists()
            is_administrador = request.user.is_staff
            is_cliente = request.user.groups.filter(name='Cliente').exists()
            is_repartidor = request.user.groups.filter(name='Repartidor').exists()
            is_empresa = request.user.groups.filter(name='Empresa').exists()
        except User.DoesNotExist:
            # Usuario huérfano: cerrar sesión
            logout(request)
            return redirect('index')

    context = {
        'productos': Producto.objects.all(),
        'inmuebles': Inmueble.objects.all(),
        'empresas': Empresa.objects.all(),
        'ofertas': Oferta.objects.all(),
        'eventos': Evento.objects.all(),
        'is_Administrador': is_administrador,
        'is_Cliente': is_cliente,
        'is_Repartidor': is_repartidor,
        'is_Empresa': is_empresa,
        'contratos': [],
        'reservas': [],
    }

    if request.user.is_authenticated:
        try:
            # Repetir la verificación para el bloque autenticado
            request.user.groups.exists()
            reservas = Reserva.objects.filter(usuario=request.user)
            context['reservas'] = reservas
            contratos = ContratoFirmado.objects.filter(usuario=request.user)
            context['contratos'] = contratos

            carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
            if carritos.exists():
                carrito = carritos.first()
            else:
                carrito = CarroCompra.objects.create(usuario=request.user)

            pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
            cuenta_total = sum(carrito.precio_total() for carrito in pagados)
            context.update({
                'pagados': pagados.exists(),
                'cuenta_total': cuenta_total,
                'carrito': carrito,
                'precio_total': carrito.precio_total(),
                'articulos': carrito.items.all(),
                'reserva_forms': [(reserva, Reserva_form(instance=reserva)) for reserva in context['reservas']],
            })
        except User.DoesNotExist:
            logout(request)
            return redirect('index')

    return render(request, 'index/index.html', context)


def buscar(request):
    context = {
        'productos': [],
        'ofertas': [],
        'eventos': [],
        'inmuebles': []
    }

    nombre = request.POST.get('buscar_nombre', '').strip()

    if nombre:
        # Filtrar solo si hay texto
        context['productos'] = Producto.objects.filter(nombreP__icontains=nombre)
        context['ofertas'] = Oferta.objects.filter(nombreO__icontains=nombre)
        context['eventos'] = Evento.objects.filter(nombreE__icontains=nombre)
        context['inmuebles'] = Inmueble.objects.filter(nombreI__icontains=nombre)

    # Pasar el término buscado para mostrarlo
    context['termino'] = nombre

    return render(request, 'index/buscar.html', context)


@login_required
def subir_pago(request, tipo, pk):

    if tipo == 'reserva':
        obj = get_object_or_404(Reserva, pk=pk, usuario=request.user)
        redirect_url = 'index'
        campo_imagen = 'pagar'
        success_msg = 'Comprobante de pago subido. Pendiente de confirmación. Puede demorar hasta 24h.'
    elif tipo == 'carro':
        obj = get_object_or_404(CarroCompra, pk=pk, usuario=request.user, cerrado=False)
        redirect_url = 'index'
        campo_imagen = 'pagar'
        success_msg = 'Pago del carrito subido. Pendiente de confirmación. Puede demorar hasta 24h.'
    else:
        raise Http404("Tipo no válido")

    if request.method == 'POST':
        imagen = request.FILES.get('imagen_pago')

        if imagen:
            setattr(obj, campo_imagen, imagen)
            obj.save()
            messages.success(request, success_msg)
        else:
            messages.error(request, 'Debes subir una imagen del comprobante.')

        return redirect(redirect_url)
    return redirect(redirect_url)


ALLOWED_MODELS = {
    'producto': Producto,
    'inmueble': Inmueble,
    'oferta': Oferta,
    'evento': Evento,
    'empresa': Empresa,
}


def agregar_media(request, content_type_str, pk):
    # Validar tipo de contenido
    if content_type_str not in ALLOWED_MODELS:
        messages.error(request, "Tipo de objeto no permitido.")
        return redirect('index')

    content_type = get_object_or_404(ContentType, model=content_type_str)
    obj = get_object_or_404(content_type.model_class(), pk=pk)
    print('object: ', obj)

    if request.method == 'POST':
        form = MediaForm(request.POST, request.FILES)
        if form.is_valid():
            media = form.save(commit=False)
            media.content_object = obj
            print('media: ', media)

            media.save()
            messages.success(request, f"Media añadida a {obj}.")
            return redirect('producto:' + content_type_str + '_detalle', pk=pk)
    else:
        form = MediaForm()

    return render(request, 'index/media.html', {
        'form': form,
        'obj': obj,
        'content_type_str': content_type_str,
    })


@login_required
def producto_add(request, empresa_pk=None):
    if request.method == "POST":
        form = Producto_form(request.POST, request.FILES)
        if form.is_valid():

            producto = form.save(commit=False)
            if empresa_pk:
                producto.empresa = Empresa.objects.get(pk=empresa_pk)
                producto.save()
                return redirect("producto:empresa_detalle", empresa_pk)

            producto.save()
            return redirect('index')
    else:
        form = Producto_form()

    context = {'form': form}

    return render(request, "producto/add_producto.html", context)


@method_decorator(login_required, name='dispatch')
class Producto_update(UpdateView):
    template_name = 'producto/actualizar_producto.html'
    queryset = Producto.objects.all()
    form_class = Producto_form

    def get_success_url(self):
        return reverse('index')


@method_decorator(login_required, name='dispatch')
class Producto_list(ListView):
    queryset = Producto.objects.all()
    template_name = 'producto/producto_list.html'


@method_decorator(login_required, name='dispatch')
class Producto_delete(DeleteView):
    template_name = 'producto/eliminar_producto.html'
    queryset = Producto.objects.all()

    def get_success_url(self):
        return reverse("index")


@method_decorator(login_required, name='dispatch')
class SubirContratoFirmadoView(View):
    def post(self, request, tipo):
        if tipo not in ['empresa', 'inmueble']:
            messages.error(request, "Tipo de contrato inválido.")
            return redirect('index')

        archivo = request.FILES.get('archivo')
        allowed_extensions = ['.jpg', '.pdf']
        if not archivo:
            messages.error(request, "Selecciona un archivo.")
        else:
            ext = os.path.splitext(archivo.name)[1].lower()
            if ext not in allowed_extensions:
                messages.error(request, "Solo se permiten PDFs o JPGs.")
            else:
                # Desactivar contratos anteriores del mismo tipo
                Contrato.objects.filter(activo=True, tipo=tipo).update(activo=False)

                # Eliminar contrato firmado anterior del mismo tipo (opcional)
                ContratoFirmado.objects.filter(usuario=request.user, tipo=tipo).delete()

                contrato_firmado = ContratoFirmado(
                    usuario=request.user,
                    archivo=archivo,
                    tipo=tipo
                )
                contrato_firmado.save()
                messages.success(request, f"Contrato de {tipo} firmado subido con éxito.")
        return redirect('producto:gestionar_contrato', tipo=tipo)


@method_decorator(login_required, name='dispatch')
class ContratoFirmadoUpdate(UpdateView):
    model = ContratoFirmado
    fields = ['archivo']
    success_url = reverse_lazy('index')  # O redirigir a una vista específica

    def get_object(self, queryset=None):
        contrato = get_object_or_404(ContratoFirmado, pk=self.kwargs['pk'])
        if contrato.usuario != self.request.user and not self.request.user.is_staff:
            raise PermissionDenied
        return contrato

    def form_valid(self, form):
        contrato = self.get_object()
        nuevo_archivo = form.cleaned_data.get('archivo')

        # Eliminar archivo anterior
        if nuevo_archivo and contrato.archivo and os.path.isfile(contrato.archivo.path):
            os.remove(contrato.archivo.path)

        # Solo staff puede cambiar estado
        if self.request.user.is_staff:
            contrato.estado = 'confirmado' in self.request.POST
            contrato.save()

        messages.success(self.request, "Contrato actualizado correctamente.")
        return redirect(self.success_url)


@login_required
def contrato_firmado_delete(request, pk):
    contrato = get_object_or_404(ContratoFirmado, pk=pk, usuario=request.user)
    pku = contrato.usuario.pk

    if request.method == 'POST':
        if contrato.archivo and os.path.isfile(contrato.archivo.path):
            os.remove(contrato.archivo.path)
        contrato.delete()
        messages.success(request, "Contrato firmado eliminado correctamente.")
    return redirect('user:details', pk=pku)


@method_decorator(login_required, name='dispatch')
class SubirContratoAdminView(View):
    def post(self, request, tipo):
        if not request.user.is_staff or tipo not in ['empresa', 'inmueble']:
            raise PermissionDenied

        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, "Selecciona un archivo.")
        else:
            ext = os.path.splitext(archivo.name)[1].lower()
            if ext not in ['.pdf', '.jpg']:
                messages.error(request, "Formato no permitido.")
            else:
                # Desactivar el contrato activo del mismo tipo
                Contrato.objects.filter(activo=True, tipo=tipo).update(activo=False)
                nuevo = Contrato(archivo=archivo, activo=True, tipo=tipo)
                nuevo.save()
                messages.success(request, f"Nuevo contrato de {tipo} subido.")
        return redirect('producto:gestionar_contrato', tipo=tipo)

    def get(self, request, tipo):
        if not request.user.is_staff:
            return redirect('index')
        return redirect('producto:gestionar_contrato', tipo=tipo)


@method_decorator(login_required, name='dispatch')
class GestionarContratoView(TemplateView):
    template_name = 'index/gestionar_contrato.html'

    def get_context_data(self, **kwargs):
        tipo = self.kwargs['tipo']  # 'empresa' o 'inmueble'
        if tipo not in ['empresa', 'inmueble']:
            raise Http404("Tipo de contrato no válido")

        context = super().get_context_data(**kwargs)
        # Contrato oficial activo del tipo
        context['contrato'] = Contrato.objects.filter(activo=True, tipo=tipo).first()
        # Contrato firmado del usuario (si existe)
        context['contrato_firmado'] = ContratoFirmado.objects.filter(
            usuario=self.request.user,
            tipo=tipo
        ).first()
        context['tipo'] = tipo
        return context


@method_decorator(login_required, name='dispatch')
class Inmueble_add(CreateView):
    template_name = 'inmueble/add_inmueble.html'
    model = Inmueble
    form_class = Inmueble_form

    def get_success_url(self):
        return reverse('index')


@method_decorator(login_required, name='dispatch')
class Inmueble_update(UpdateView):
    template_name = 'inmueble/actualizar_inmueble.html'
    queryset = Inmueble.objects.all()
    form_class = Inmueble_form

    def get_success_url(self):
        return reverse('index')


@method_decorator(login_required, name='dispatch')
class Inmueble_list(ListView):
    queryset = Inmueble.objects.all()
    template_name = 'inmueble/inmueble_list.html'


@method_decorator(login_required, name='dispatch')
class Inmueble_delete(DeleteView):
    template_name = 'inmueble/eliminar_inmueble.html'
    queryset = Inmueble.objects.all()

    def get_success_url(self):
        return reverse("index")


@method_decorator(login_required, name='dispatch')
class Empresa_add(CreateView):
    model = Empresa
    form_class = Empresa_form
    template_name = 'empresa/add_empresa.html'
    success_url = reverse_lazy('index')

    def form_valid(self, form):
        return super().form_valid(form)


@login_required
def empresa_update(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == "POST":
        form = Empresa_form(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save()
            return redirect('producto:empresa_detalle', empresa_pk=empresa.pk)
    else:
        form = Empresa_form(instance=empresa)

    return render(request, 'empresa/actualizar_empresa.html', {
        'form': form,
        'empresa': empresa
    })


@method_decorator(login_required, name='dispatch')
class Empresa_delete(DeleteView):
    template_name = 'empresa/eliminar_empresa.html'
    queryset = Empresa.objects.all()

    def get_success_url(self):
        return reverse("index")


@login_required
def oferta_add(request, empresa_pk=None):
    if request.method == "POST":
        form = Oferta_form(request.POST, request.FILES)
        if form.is_valid():

            oferta = form.save(commit=False)
            if empresa_pk:
                oferta.empresa = Empresa.objects.get(pk=empresa_pk)
                oferta.save()
                return redirect("producto:empresa_detalle", empresa_pk)

            oferta.save()
            return redirect('index')
    else:
        form = Oferta_form()

    context = {'form': form}

    return render(request, "oferta/add_oferta.html", context)


@method_decorator(login_required, name='dispatch')
class Oferta_update(UpdateView):
    template_name = 'oferta/actualizar_oferta.html'
    queryset = Oferta.objects.all()
    form_class = Oferta_form

    def get_success_url(self):
        return reverse('index')


@method_decorator(login_required, name='dispatch')
class Oferta_list(ListView):
    queryset = Oferta.objects.all()
    template_name = 'oferta/oferta_list.html'


@method_decorator(login_required, name='dispatch')
class Oferta_delete(DeleteView):
    template_name = 'oferta/eliminar_oferta.html'
    queryset = Oferta.objects.all()

    def get_success_url(self):
        return reverse("index")


@login_required
def evento_add(request, empresa_pk=None):
    if request.method == "POST":
        form = Evento_form(request.POST, request.FILES)
        if form.is_valid():

            evento = form.save(commit=False)
            if empresa_pk:
                evento.empresa = Empresa.objects.get(pk=empresa_pk)
                evento.save()
                return redirect("producto:empresa_detalle", empresa_pk)

            evento.save()
            return redirect('index')
    else:
        form = Evento_form()

    context = {'form': form}

    return render(request, "evento/add_evento.html", context)


@method_decorator(login_required, name='dispatch')
class Evento_update(UpdateView):
    template_name = 'evento/actualizar_evento.html'
    queryset = Evento.objects.all()
    form_class = Evento_form

    def get_success_url(self):
        return reverse('index')


@method_decorator(login_required, name='dispatch')
class Evento_list(ListView):
    queryset = Evento.objects.all()
    template_name = 'evento/evento_list.html'


@method_decorator(login_required, name='dispatch')
class Evento_delete(DeleteView):
    template_name = 'evento/eliminar_evento.html'
    queryset = Evento.objects.all()

    def get_success_url(self):
        return reverse("index")


@method_decorator(login_required, name='dispatch')
class Reserva_add(CreateView):
    template_name = 'Evento/add_reserva.html'
    form_class = Reserva_form
    success_url = reverse_lazy('index')  # Cambia si quieres redirigir a otra página

    def form_valid(self, form):
        evento_id = self.kwargs.get('pk')
        evento = get_object_or_404(Evento, pk=evento_id)

        # Asignar usuario y evento
        form.instance.usuario = self.request.user
        form.instance.evento = evento

        # Si el usuario no es staff ni superusuario, forzar estado=False
        if not self.request.user.is_staff and not self.request.user.is_superuser:
            form.instance.estado = False

        messages.success(
            self.request,
            f"✅ Reserva creada con éxito para '{evento.nombreE}'."
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request,
            "❌ Por favor corrige los errores del formulario."
        )
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        evento_id = self.kwargs.get('pk')
        context['evento'] = get_object_or_404(Evento, pk=evento_id)
        return context


@login_required
def editar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk, usuario=request.user)
    pku = reserva.usuario.pk

    # No se puede editar si ya está confirmada (solo el admin puede saltarse esta regla)
    if reserva.estado and not request.user.is_staff:
        messages.warning(
            request,
            "⚠️ No puedes editar una reserva que ya ha sido confirmada."
        )
        return redirect('index')

    if request.method == 'POST':
        # === Actualizar num_personas (solo si no es admin o si se permite) ===
        num_personas = request.POST.get('num_personas')

        if not request.user.is_staff:  # Solo los usuarios normales necesitan validar cambios de personas
            if not num_personas or not num_personas.isdigit():
                messages.error(request, "❌ La cantidad de personas debe ser un número válido.")
            else:
                num_personas = int(num_personas)
                if num_personas < 1:
                    messages.error(request, "❌ Debes reservar al menos 1 persona.")
                elif num_personas > 10:
                    messages.error(request, "❌ Máximo 10 personas por reserva.")
                else:
                    reserva.num_personas = num_personas
                    reserva.save()
                    messages.success(
                        request,
                        f"✅ Reserva actualizada: {num_personas} persona(s) para '{reserva.evento.nombreE}'."
                    )
        else:
            # Admin puede cambiar num_personas directamente sin límite (opcional)
            if num_personas and num_personas.isdigit():
                reserva.num_personas = int(num_personas)
                reserva.save()

        # === Actualizar estado (solo si el usuario es staff) ===
        if request.user.is_staff:
            nuevo_estado = request.POST.get('estado')
            estado_bool = nuevo_estado == 'on'  # Checkbox: presente = True
            if reserva.estado != estado_bool:
                reserva.estado = estado_bool
                reserva.save()
                messages.success(
                    request,
                    f"✅ Estado de la reserva actualizado: {'Confirmada' if estado_bool else 'Pendiente'}."
                )

        # Redirigir según rol
        if request.user.is_staff:
            return redirect('user:details', pk=pku)
        else:
            return redirect('index')

    return redirect('index')


@login_required
def eliminar_reserva(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk, usuario=request.user)
    pku = reserva.usuario.pk

    if reserva.estado:
        messages.warning(
            request,
            "No puedes eliminar una reserva ya confirmada."
        )
    else:
        evento_nombre = reserva.evento.nombreE
        reserva.delete()
        messages.success(
            request,
            f"Reserva eliminada con éxito para '{evento_nombre}'."
        )

    if request.user.is_staff:
        return redirect('user:details', pk=pku)
    else:
        return redirect('index')


@method_decorator(login_required, name='dispatch')
class Reserva_list(ListView):
    queryset = Reserva.objects.all()
    template_name = 'evento/reserva_list.html'


@method_decorator(login_required, name='dispatch')
class Reserva_delete(DeleteView):
    template_name = 'evento/eliminar_reserva.html'
    queryset = Reserva.objects.all()

    def get_success_url(self):
        return reverse("index")


def producto_detalle(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    valoraciones = producto.valoraciones.all().order_by('-fecha')

    # Contexto base
    context = {
        'producto': producto,
        'valoraciones': valoraciones,
    }

    # Manejo de carrito y datos relacionados SOLO si está autenticado
    if request.user.is_authenticated:
        carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
        if carritos.exists():
            carrito = carritos.first()
        else:
            carrito = CarroCompra.objects.create(usuario=request.user)

        pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
        cuenta_total = sum(c.precio_total() for c in pagados)

        content_type = ContentType.objects.get_for_model(Producto)
        medias = Media.objects.filter(
            content_type=content_type,
            object_id=producto.pk
        ).order_by('orden', 'pk')

        context.update({
            'carrito': carrito,
            'cuenta_total': cuenta_total,
            'pagados': pagados.exists(),
            'precio_total': carrito.precio_total(),
            'articulos': carrito.items.all(),
            'medias': medias,
        })
    else:
        # Si no está autenticado, no hay carrito ni medias relacionadas con usuario
        # Pero aún queremos mostrar las medias del producto (no dependen del usuario)
        content_type = ContentType.objects.get_for_model(Producto)
        medias = Media.objects.filter(
            content_type=content_type,
            object_id=producto.pk
        ).order_by('orden', 'pk')
        context['medias'] = medias

    # Manejo del formulario de valoración (solo POST y autenticado)
    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, 'Debes iniciar sesión para enviar una valoración.')
            return redirect('login')  # o 'user:login' según tu configuración

        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '').strip()
        if valor:  # asegúrate de que se haya enviado un valor
            Valoracion.objects.create(
                usuario=request.user,
                producto=producto,
                valor=valor,
                comentario=comentario
            )
            messages.success(request, 'Gracias por tu valoración.')
        return redirect('producto:producto_detalle', pk=pk)

    return render(request, 'detalles/producto_detalle.html', context)


def inmueble_detalle(request, pk):
    inmueble = get_object_or_404(Inmueble, pk=pk)
    valoraciones = inmueble.valoraciones.all().order_by('-fecha')

    context = {
        'inmueble': inmueble,
        'valoraciones': valoraciones,
    }

    if request.user.is_authenticated:
        carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
        carrito = carritos.first() if carritos.exists() else CarroCompra.objects.create(usuario=request.user)

        pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
        cuenta_total = sum(c.precio_total() for c in pagados)

        content_type = ContentType.objects.get_for_model(Inmueble)
        medias = Media.objects.filter(content_type=content_type, object_id=inmueble.pk).order_by('orden', 'pk')

        context.update({
            'carrito': carrito,
            'cuenta_total': cuenta_total,
            'pagados': pagados.exists(),
            'precio_total': carrito.precio_total(),
            'articulos': carrito.items.all(),
            'medias': medias,
        })
    else:
        content_type = ContentType.objects.get_for_model(Inmueble)
        medias = Media.objects.filter(content_type=content_type, object_id=inmueble.pk).order_by('orden', 'pk')
        context['medias'] = medias

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, 'Debes iniciar sesión para enviar una valoración.')
            return redirect('login')

        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '').strip()
        if valor:
            Valoracion.objects.create(
                usuario=request.user,
                inmueble=inmueble,
                valor=valor,
                comentario=comentario
            )
            messages.success(request, 'Gracias por tu opinión.')
        return redirect('producto:inmueble_detalle', pk=pk)

    return render(request, 'detalles/inmueble_detalle.html', context)


def oferta_detalle(request, pk):
    oferta = get_object_or_404(Oferta, pk=pk)
    valoraciones = oferta.valoraciones.all().order_by('-fecha')

    context = {
        'oferta': oferta,
        'valoraciones': valoraciones,
    }

    if request.user.is_authenticated:
        carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
        carrito = carritos.first() if carritos.exists() else CarroCompra.objects.create(usuario=request.user)

        pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
        cuenta_total = sum(c.precio_total() for c in pagados)

        content_type = ContentType.objects.get_for_model(Oferta)
        medias = Media.objects.filter(content_type=content_type, object_id=oferta.pk).order_by('orden', 'pk')

        context.update({
            'carrito': carrito,
            'cuenta_total': cuenta_total,
            'pagados': pagados.exists(),
            'precio_total': carrito.precio_total(),
            'articulos': carrito.items.all(),
            'medias': medias,
        })
    else:
        content_type = ContentType.objects.get_for_model(Oferta)
        medias = Media.objects.filter(content_type=content_type, object_id=oferta.pk).order_by('orden', 'pk')
        context['medias'] = medias

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, 'Debes iniciar sesión para enviar una valoración.')
            return redirect('login')

        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '').strip()
        if valor:
            Valoracion.objects.create(
                usuario=request.user,
                oferta=oferta,
                valor=valor,
                comentario=comentario
            )
            messages.success(request, '¡Tu opinión ha sido enviada!')
        return redirect('producto:oferta_detalle', pk=pk)

    return render(request, 'detalles/oferta_detalle.html', context)


def evento_detalle(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    valoraciones = evento.valoraciones.all().order_by('-fecha')

    context = {
        'evento': evento,
        'valoraciones': valoraciones,
    }

    if request.user.is_authenticated:
        carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
        carrito = carritos.first() if carritos.exists() else CarroCompra.objects.create(usuario=request.user)

        pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
        cuenta_total = sum(c.precio_total() for c in pagados)

        content_type = ContentType.objects.get_for_model(Evento)
        medias = Media.objects.filter(content_type=content_type, object_id=evento.pk).order_by('orden', 'pk')

        context.update({
            'carrito': carrito,
            'cuenta_total': cuenta_total,
            'pagados': pagados.exists(),
            'precio_total': carrito.precio_total(),
            'articulos': carrito.items.all(),
            'medias': medias,
        })
    else:
        content_type = ContentType.objects.get_for_model(Evento)
        medias = Media.objects.filter(content_type=content_type, object_id=evento.pk).order_by('orden', 'pk')
        context['medias'] = medias

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, 'Debes iniciar sesión para enviar una valoración.')
            return redirect('login')  # o 'user:login'

        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '').strip()
        if valor:
            Valoracion.objects.create(
                usuario=request.user,
                evento=evento,
                valor=valor,
                comentario=comentario
            )
            messages.success(request, 'Gracias por tu feedback.')
        return redirect('producto:evento_detalle', pk=pk)

    return render(request, 'detalles/evento_detalle.html', context)


def empresa_detalle(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    valoraciones = empresa.valoraciones.all().order_by('-fecha')

    context = {
        'empresa': empresa,
        'eventos': empresa.evento.all(),
        'ofertas': empresa.oferta.all(),
        'productos': empresa.producto.all(),
        'valoraciones': valoraciones,
    }

    if request.user.is_authenticated:
        carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False, cerrado=False)
        carrito = carritos.first() if carritos.exists() else CarroCompra.objects.create(usuario=request.user)

        pagados = CarroCompra.objects.filter(usuario=request.user, pagado=True, cerrado=False)
        cuenta_total = sum(c.precio_total() for c in pagados)

        content_type = ContentType.objects.get_for_model(Empresa)
        medias = Media.objects.filter(content_type=content_type, object_id=empresa.pk).order_by('orden', 'pk')

        context.update({
            'carrito': carrito,
            'cuenta_total': cuenta_total,
            'pagados': pagados.exists(),
            'precio_total': carrito.precio_total(),
            'articulos': carrito.items.all(),
            'medias': medias,
        })
    else:
        content_type = ContentType.objects.get_for_model(Empresa)
        medias = Media.objects.filter(content_type=content_type, object_id=empresa.pk).order_by('orden', 'pk')
        context['medias'] = medias

    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.warning(request, 'Debes iniciar sesión para enviar una valoración.')
            return redirect('login')

        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '').strip()
        if valor:
            Valoracion.objects.create(
                usuario=request.user,
                empresa=empresa,
                valor=valor,
                comentario=comentario
            )
            messages.success(request, 'Gracias por tu feedback.')
        return redirect('producto:empresa_detalle', pk=pk)  # ✅ Corregido

    return render(request, 'detalles/empresa_detalle.html', context)


def export_venta_to_excel(carrito):
    # Establecer el nombre del archivo Excel con la fecha de envío del carrito
    file_name = f"ventas {carrito.fecha_envio.strftime('%m-%y')}.xlsx"

    # Establecer el directorio donde se guardará el archivo Excel
    dir_name = 'registros de ventas'

    # Crear el directorio si no existe
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    # Establecer la ruta completa del archivo Excel
    file_path = os.path.join(dir_name, file_name)

    # Verificar si el archivo Excel existe
    if os.path.exists(file_path):
        # Cargar el archivo Excel existente
        wb = load_workbook(file_path)
    else:
        # Crear un nuevo archivo Excel si no existe
        wb = openpyxl.Workbook()
        ws = wb.active

        # Establecer los títulos de las columnas
        ws['A1'] = 'Cliente'
        ws['B1'] = 'Fecha'
        ws['C1'] = 'Orden'
        ws['D1'] = 'Precio'

    # Seleccionar la hoja de trabajo activa
    ws = wb.active

    # Leer la última fila escrita en el archivo Excel
    last_row = ws.max_row

    # Establecer la fila inicial para escribir nuevos datos
    row_num = last_row + 1 if last_row is not None else 1

    # crear un str orden con todos los articulos dentro del carrito separados por coma
    orden = ', '.join(str(item.cantidad) + ' ' + str(item) for item in carrito.items.all()).rstrip(', ')

    # Escribir la venta en el archivo Excel
    ws[f'A{row_num}'] = carrito.usuario.username
    ws[f'B{row_num}'] = carrito.fecha_envio.replace(tzinfo=None)
    ws[f'C{row_num}'] = orden
    ws[f'D{row_num}'] = carrito.precio_total()
    row_num += 1

    # Guardar los cambios en el archivo Excel
    wb.save(file_path)


@login_required
def vistarepartidor(request):
    # Verificar que el usuario pertenezca al grupo "Repartidor"
    if not request.user.groups.filter(name='Repartidor').exists():
        messages.error(request, "⚠️ No tienes permiso para acceder aquí.")
        return redirect('index')

    # Obtener carritos: pagados, con fecha de envío, y no cerrados
    carritos = CarroCompra.objects.filter(
        pagado=True,
        fecha_envio__isnull=False,
        cerrado=False
    ).select_related('usuario').prefetch_related('items')

    if request.method == 'POST':
        ids_marcados = []

        # Recoger todos los checkboxes marcados
        for key, value in request.POST.items():
            if key.startswith('check-'):
                if value.strip():
                    try:
                        carrito_id = int(value)
                        ids_marcados.append(carrito_id)
                    except (ValueError, TypeError):
                        continue

        # Actualizar los carritos marcados como entregados
        count = CarroCompra.objects.filter(
            id__in=ids_marcados,
            pagado=True,
            cerrado=False
        ).update(cerrado=True)

        if count > 0:
            messages.success(request, f"✅ {count} entrega(s) marcada(s) como completada(s).")
        else:
            messages.warning(request, "⚠️ No se actualizó ninguna entrega.")

        return redirect('producto:repartidor')

    return render(request, 'user/vista_repartidor.html', {
        'carritos': carritos
    })


@login_required
def marcar_pago(request, pk):
    if not request.user.is_staff:
        messages.error(request, "No tienes permiso.")
        return redirect('user:list')

    carrito = get_object_or_404(CarroCompra, pk=pk)

    if request.method == 'POST':
        carrito.pagado = True
        if not carrito.fecha_pago:
            carrito.fecha_pago = timezone.now()
        if not carrito.fecha_envio:
            carrito.fecha_envio = timezone.now()
        carrito.save()

        export_venta_to_excel(carrito)

        messages.success(request, f"✅ Pago confirmado para {carrito.usuario.username}.")
    return redirect('user:list')


@login_required
def pedir_cuenta(request):
    suma = 0
    carritos = request.user.carritos.filter(pagado=True, cerrado=False)
    for carrito in carritos:
        carrito.cerrado = True
        suma += carrito.precio_total()
        carrito.save()
    if carritos.last().cerrado:
        messages.success(request, f'Su pedido fue entregado con éxito.')
    return redirect('index')


@login_required
def add_producto_carrito(request, pk):
    producto = Producto.objects.get(pk=pk)
    carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False)
    if not carritos:
        carrito = CarroCompra.objects.create(usuario=request.user)
    else:
        carrito = carritos[0]

    carrito_item, creado = carrito.items.get_or_create(producto=producto)
    carrito_item.cantidad += 1
    carrito_item.save()

    if carrito_item:
        messages.success(request, f'Artículo "{producto.nombreP}" añadido al carrito.')

    return redirect('index')


@login_required
def add_oferta_carrito(request, pk):
    oferta = Oferta.objects.get(pk=pk)
    carritos = CarroCompra.objects.filter(usuario=request.user, pagado=False)
    if not carritos:
        carrito = CarroCompra.objects.create(usuario=request.user)
    else:
        carrito = carritos[0]

    carrito_item, creado = carrito.items.get_or_create(oferta=oferta)
    carrito_item.cantidad += 1
    carrito_item.save()

    if carrito_item:
        messages.success(request, f'Artículo "{oferta.nombreO}" añadido al carrito.')

    return redirect('index')


@login_required
def ajustar_carrito_item(request, pk, tipo):
    carrito_item = ArticuloCarro.objects.get(pk=pk)

    if tipo == "up":
        carrito_item.cantidad += 1
    elif tipo == "down":
        carrito_item.cantidad -= 1
    carrito_item.save()

    return redirect('index')


@login_required
def eliminar_carrito_item(request, pk):
    carrito_item = ArticuloCarro.objects.get(pk=pk)
    carrito_item.delete()

    return redirect('index')


@login_required
def crear_valoracion(request, pk):
    inmueble = get_object_or_404(Inmueble, pk=pk)

    if request.method == 'POST':
        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '')

        if valor and valor.isdigit() and 1 <= int(valor) <= 5:
            Valoracion.objects.create(
                usuario=request.user,
                inmueble=inmueble,
                valor=int(valor),
                comentario=comentario
            )
            messages.success(request, "✅ Gracias por tu opinión.")
        else:
            messages.error(request, "❌ Selecciona una calificación válida.")

    return redirect('index')  # O a la lista si prefieres


@login_required
def editar_valoracion(request, pk):
    val = get_object_or_404(Valoracion, pk=pk, usuario=request.user)

    if request.method == 'POST':
        valor = request.POST.get('valor')
        comentario = request.POST.get('comentario', '')

        if valor and valor.isdigit() and 1 <= int(valor) <= 5:
            val.valor = int(valor)
            val.comentario = comentario
            val.save()
            messages.success(request, "✅ Valoración actualizada con éxito.")
        else:
            messages.error(request, "❌ Selecciona una calificación válida.")

    return redirect('producto:inmueble_detalle',
                    pk=val.inmueble.pk
                    if val.inmueble else val.producto.pk
                    if val.producto else val.evento.pk
                    if val.evento else val.oferta.pk)


@login_required
def eliminar_valoracion(request, pk):
    val = get_object_or_404(Valoracion, pk=pk, usuario=request.user)

    # Guardamos el objeto relacionado antes de eliminar
    inmueble_pk = val.inmueble.pk if val.inmueble else None
    producto_pk = val.producto.pk if val.producto else None
    evento_pk = val.evento.pk if val.evento else None
    oferta_pk = val.oferta.pk if val.oferta else None

    val.delete()
    messages.success(request, "🗑️ Valoración eliminada con éxito.")

    # Redirigir al detalle correspondiente
    if inmueble_pk:
        return redirect('producto:inmueble_detalle', pk=inmueble_pk)
    elif producto_pk:
        return redirect('producto:producto_detalle', pk=producto_pk)
    elif evento_pk:
        return redirect('producto:evento_detalle', pk=evento_pk)
    elif oferta_pk:
        return redirect('producto:oferta_detalle', pk=oferta_pk)
    else:
        return redirect('index')


@login_required
def limpiar_reservas(request):
    # Solo accesible para administradores
    if not request.user.is_staff:
        return HttpResponse("Acceso denegado", status=403)

    # Calcular la fecha límite: 24 horas atrás
    limite = datetime.now() - timedelta(hours=24)

    # Obtener eventos que ya pasaron hace más de 24h
    eventos_pasados = Evento.objects.filter(fechahora__lt=limite)

    # Contar cuántas reservas se eliminarán
    reservas = Reserva.objects.filter(evento__in=eventos_pasados)
    count = reservas.count()

    # Eliminarlas
    reservas.delete()

    # Mensaje de éxito
    if count > 0:
        messages.success(request, f"✅ Se eliminaron {count} reservas caducadas.")
    else:
        messages.info(request, "✅ No había reservas caducadas para eliminar.")

    return redirect('user:list')


